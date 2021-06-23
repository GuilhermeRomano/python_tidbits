'''
Program to ingest student data and produce individual reports in
a spreadsheet
'''

import openpyxl
from tkinter import *
import tkinter.filedialog
import os
import pprint
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
from openpyxl.utils import get_column_letter
import pandas as pd
import datetime
import numpy as np
from statistics import mean, stdev

# hide tkinter window
Tk().withdraw()


def select_workbook(window_title):
    filetypes = (
        ("Excel Workbook", "*.xlsx"),
        ("Excel Macro-Enabled Workbook (code)", "*.xlsm"),
    )

    filename = tkinter.filedialog.askopenfilename(
        title=window_title,
        initialdir="C:/Users/Guilherme/Documents/MEGA/My Files/Projects/Projetos Externos/Prof Trindade/Block Dedication",
        filetypes=filetypes)

    if filename == "":
        print("Please, select a workbook")
        exit()
    else:
        return filename


def change_directory(destination):

    try:
        # Change the current working Directory
        os.chdir(destination)
    except OSError:
        print("Can\'t change the Current Working Directory")


def as_text(value):
    if value is None:
        return ""
    return str(value)


def column_fit_properly(worksheet):
    for column_cells in worksheet.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells)
        worksheet.column_dimensions[get_column_letter(
            column_cells[0].column)].width = length


# Select control workbook
control_location = select_workbook(
    window_title="Select Control Workbook"
)
control_book = openpyxl.load_workbook(
    control_location, read_only=False, keep_vba=True)
gabarito_sheet = control_book["Gabarito"]

change_directory(os.path.dirname(control_location))

if gabarito_sheet.max_row == 1:
    print("No activities in template!")
    exit()

atividade_list = []
time_delta_dict = {}
for atividade in range(2, gabarito_sheet.max_row+1):
    atividade_title = gabarito_sheet.cell(atividade, 1).value
    time_delta = gabarito_sheet.cell(atividade, 2).value

    if atividade_title != "Dedicação Geral Independente da Atividade estar no Gabarito":
        atividade_list.append(atividade_title)

    if time_delta == 0 or time_delta == "" or time_delta is None:
        print("Operation Cancelled!\nInsert time delta")
        exit()
    else:
        time_delta_dict[atividade_title] = pd.to_timedelta(
            time_delta, unit="minute")

time_delta_general = time_delta_dict["Dedicação Geral Independente da Atividade estar no Gabarito"]

# Limpar sheets de atividades passadas
for sheet in control_book.sheetnames:
    if sheet != "Gabarito":
        del control_book[sheet]

# Ingerir logs
source_location = select_workbook(window_title="Select Activity Logs")
source_book = openpyxl.load_workbook(source_location)
source_sheet_name = source_book.sheetnames[0]

source_df = pd.read_excel(
    source_location, sheet_name=source_sheet_name, parse_dates=["Hora"],
)
# filtrar colunas uteis
atividades_df = source_df[["Hora", "Nome completo", "Contexto do Evento"]]

# Organizar em ordem crescente
atividades_df = atividades_df.sort_values(
    by=["Nome completo", "Contexto do Evento", "Hora"])

# Criar a lista de alunos
student_list = list(set(atividades_df["Nome completo"].to_numpy()))
student_list.sort()

# Styles
# Header
header_style = NamedStyle(name="Header Style")
header_style.font = Font(bold=True,
                         color='00FFFFFF')
header_style.fill = PatternFill(fill_type="solid",
                                start_color='FF000000',
                                )

if header_style.name not in control_book.style_names:
    control_book.add_named_style(header_style)

# make sure our dates will be stored prrperly
control_book.iso_dates = True
control_book.epoch = openpyxl.utils.datetime.CALENDAR_MAC_1904
# Preparar dicionarios
activity_sum_dict = {}
activity_desvp_dict = {}
for activity in range(len(atividade_list)):
    activity_name = atividade_list[activity]
    activity_sum_dict.setdefault(activity_name, pd.Timedelta(seconds=0))
    activity_desvp_dict.setdefault(activity_name, 0)

dedication_specific_dict = {}
dedication_general_dict = {}
for student in range(len(student_list)):
    student_name = student_list[student]
    dedication_specific_dict.setdefault(student_name, pd.Timedelta(seconds=0))
    dedication_general_dict.setdefault(student_name, pd.Timedelta(seconds=0))

# Para cada aluno, preencher os dados de cada atividade
for student in range(len(student_list)):
    student_name = student_list[student]
    student_sheet = control_book.create_sheet(title=student_name[:31])

    # headers
    student_sheet.cell(1, 1).value = "Atividade"
    student_sheet.cell(1, 2).value = "Dedicação (h:mm)"

    # get a slice of an student
    student_df = atividades_df[atividades_df["Nome completo"] == student_name][
        ["Hora", "Contexto do Evento"]]
    student_df = student_df.sort_values(by=["Contexto do Evento", "Hora"])
    student_df["activity dt"] = student_df.groupby(
        ["Contexto do Evento"])["Hora"].apply(lambda x: x.diff())
    student_df["activity dt"].fillna(
        value=pd.Timedelta(seconds=0), inplace=True)

    # initialize standard values
    dedicacao_especifica = pd.Timedelta(seconds=0)
    dedicacao_especifica_total = pd.Timedelta(seconds=0)
    dedicacao_geral = pd.Timedelta(seconds=0)

    # Loop para dedicacao especifica
    for activity in range(len(atividade_list)):
        activity_rows = activity+2  # variavel para facilitar o loop
        activity_name = atividade_list[activity]
        time_delta = time_delta_dict[activity_name]
        # get the slice for an activity
        activity_df = student_df[(student_df["Contexto do Evento"] == activity_name) & (student_df["activity dt"] <= time_delta)][
            ["Hora", "activity dt"]]

        dedicacao_especifica = activity_df["activity dt"].sum()
        activity_sum_dict[activity_name] += dedicacao_especifica
        dedicacao_especifica_total = dedicacao_especifica_total+dedicacao_especifica
        student_sheet.cell(activity_rows, 1).value = atividade_list[activity]
        student_sheet.cell(activity_rows, 2).value = dedicacao_especifica
        student_sheet.cell(activity_rows, 2).number_format = "h:mm"

    student_df = student_df.sort_values(by=["Hora", "Contexto do Evento"])
    student_df["overall dt"] = (
        student_df["Hora"]-student_df["Hora"].shift()).fillna(pd.Timedelta(seconds=0))
    dedicacao_geral = student_df[student_df["overall dt"] <= time_delta_general][
        "overall dt"].sum()

    # headers
    student_sheet.cell(activity_rows+1, 1).value = "Total das atividades acima"
    student_sheet.cell(
        activity_rows+2, 1).value = "Dedicação Geral Independente da Atividade estar no Gabarito"
    # values
    student_sheet.cell(activity_rows+1, 2).value = dedicacao_especifica_total
    student_sheet.cell(activity_rows+2, 2).value = dedicacao_geral
    # fill dictionaries
    dedication_specific_dict[student_name] += dedicacao_especifica_total
    dedication_general_dict[student_name] += dedicacao_geral
    # formatting
    student_sheet.cell(1, 1).style = "Header Style"
    student_sheet.cell(1, 2).style = "Header Style"
    student_sheet.cell(activity_rows+1, 2).number_format = "h:mm"
    student_sheet.cell(activity_rows+2, 2).number_format = "h:mm"

    column_fit_properly(student_sheet)

# Preencher a tabela de resultados

summary_sheet = control_book.create_sheet(title="Resultado", index=1)

summary_headers = ["Aluno",
                   "Tempo Dedicado Total Específico (h:mm)",
                   "Tempo Dedicado Total Relativo (%)",
                   "Tempo Dedicado Geral Absoluto (h:mm)",
                   "Tempo Dedicado Geral Relativo (%)"]
for column in summary_sheet.columns:
    summary_sheet.cell(1, column).value = summary_headers[column-1]
    summary_sheet.cell(1, column).style = "Header Style"

dedication_specific_max = max(dedication_specific_dict.values())
dedication_general_max = max(dedication_general_dict.values())

for student in range(len(student_list)):
    student_rows = student+2
    student_name = student_list[student]
    # values
    summary_sheet.cell(student_rows, 1).value = student_name
    summary_sheet.cell(
        student_rows, 2).value = dedication_specific_dict[student_name]
    summary_sheet.cell(student_rows, 3).value = dedication_specific_dict[student_name].total_seconds(
    ) / dedication_specific_max.total_seconds()
    summary_sheet.cell(
        student_rows, 4).value = dedication_general_dict[student_name]
    summary_sheet.cell(student_rows, 5).value = dedication_general_dict[student_name].total_seconds(
    ) / dedication_general_max.total_seconds()

    # formatting
    summary_sheet.cell(student_rows, 2).number_format = "h:mm"
    summary_sheet.cell(student_rows, 3).style = "Percent"
    summary_sheet.cell(student_rows, 3).number_format = "0.00%"
    summary_sheet.cell(student_rows, 4).number_format = "h:mm"
    summary_sheet.cell(student_rows, 5).style = "Percent"
    summary_sheet.cell(student_rows, 5).number_format = "0.00%"
# sdfsdf
# Calcular Media e Desvio Padrao
for activity in range(len(atividade_list)):
    activity_rows = activity + 2
    activity_name = atividade_list[activity]
    # values
    average = activity_sum_dict[activity_name].total_seconds() / \
        (len(atividade_list)+1)
    gabarito_sheet.cell(activity_rows, 3).value = average

    for sheet in control_book.sheetnames[1:len(control_book.sheetnames)]:
        student_sheet = control_book[sheet]
        student_max_row = student_sheet.max_row
        for row in range(2, student_sheet.max_row+1):
            search_activity = student_sheet.cell(row, 1).value
            if search_activity == activity_name:
                sample = student_sheet.cell(row, 2).value.total_seconds()
                activity_desvp_dict[activity_name] += (sample - average)**2

    gabarito_sheet.cell(activity_rows, 3).value = (
        activity_desvp_dict[activity_name] / (len(atividade_list)+1))**(1/2)

    # formatting
    gabarito_sheet.cell(activity_rows, 3).number_format = "h:mm"
    gabarito_sheet.cell(activity_rows, 4).number_format = "h:mm"

# Totals
# get values as lists
dedication_general_list = dedication_general_dict.values()

dedication_general_sec_list = []
# convert that to seconds
for value in dedication_general_list:
    value_seconds = value.total_seconds()
    dedication_general_sec_list.append(value_seconds)

gabarito_sheet.cell(gabarito_sheet.max_row, 3).value = mean(
    dedication_general_sec_list)

gabarito_sheet.cell(gabarito_sheet.max_row, 3).value = stdev(
    dedication_general_sec_list)

# formatting
gabarito_sheet.cell(gabarito_sheet.max_row, 4).number_format = "h:mm"
gabarito_sheet.cell(gabarito_sheet.max_row, 3).number_format = "h:mm"

column_fit_properly(summary_sheet)

# save and finish
control_book.save(filename=os.path.basename(control_location))
print("\nUpdated\n")
